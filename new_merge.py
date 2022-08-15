import openpyxl as opxl
import glob
import xlsxwriter as xs
import tkinter as tk
from tkinter import filedialog as fd
import pandas as pd
import time

# ask_dir asks user to open a folder where all the excel output is located
# and returns the folder's location

def ask_dir():
    root = tk.Tk()
    root.title('Open File Dialog')
    root.geometry('0x0')
    root.resizable(False, False)
    path = fd.askdirectory(title = 'Open a file', initialdir = '/', mustexist = True)
    root.destroy()
    return path

# find_length returns the length of the file as dataframe + 2 for formatting purposes

def find_length(file):
    df = pd.read_excel(file)
    length = len(df)
    # print(df)
    # print(length)
    return length + 2

def main():
    path = ask_dir()

    file_list = glob.glob(path + '/*.xlsx')
    print("Select a folder where Weekly Report will output.")
    time.sleep(2)
    save_path = ask_dir()
    wb = xs.Workbook(save_path + "/Weekly_Report.xlsx")
    
    ws = wb.add_worksheet("Report")
    k = 1

    # this triple inbedded loop could be optimized in the future as it obtains a quite inefficient runtime
    # this triple inbedded loop loops through each output file and each output file's data

    # while this is a triple loop, it obtains a O(n^2) run time since the inner loop is constant value

    for file in file_list:
        wb_obj = opxl.load_workbook(file)
        s_obj = wb_obj.active
        length = find_length(file)
        for i in range(1, length):
            for j in range(1,3):
                cell = s_obj.cell(row = i, column = j)
                value = cell.value
                ws.write(k, j, value)
            k += 1
    wb.close()


